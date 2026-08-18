import sys
import os
import io
from pathlib import Path
from typing import Tuple, List, Optional
from datetime import datetime

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Charger Fault Executive Dashboard",
    layout="wide",
    initial_sidebar_state="expanded"
)

DEFAULT_PATH = sys.argv[1] if len(sys.argv) > 1 else os.environ.get("FAULT_SHEET_PATH", "")
DATE_COLS = ["Issue Date", "Resolution Date", "Restoration Date"]
CORPORATE_PALETTE = ["#1E3A8A", "#0284C7", "#0D9488", "#F59E0B", "#E11D48", "#8B5CF6", "#64748B"]


def inject_custom_css():
    st.markdown("""
    <style>
        .main .block-container {
            padding-top: 1.2rem;
            padding-bottom: 2rem;
            max-width: 96%;
        }
        .header-banner {
            background: linear-gradient(135deg, #0F172A 0%, #1E3A8A 100%);
            color: #FFFFFF;
            padding: 1.5rem 2rem;
            border-radius: 12px;
            margin-bottom: 1.5rem;
            box-shadow: 0 4px 15px rgba(15, 23, 42, 0.12);
        }
        .header-title {
            font-size: 2.1rem;
            font-weight: 700;
            margin: 0;
            color: #FFFFFF;
            letter-spacing: -0.5px;
        }
        .header-subtitle {
            font-size: 0.95rem;
            color: #94A3B8;
            margin-top: 0.3rem;
            margin-bottom: 0.5rem;
        }
        .header-badge {
            background: rgba(56, 189, 248, 0.15);
            border: 1px solid rgba(56, 189, 248, 0.4);
            color: #38BDF8;
            padding: 0.3rem 0.85rem;
            border-radius: 20px;
            font-size: 0.82rem;
            font-weight: 600;
            display: inline-block;
        }
        div[data-testid="stMetric"] {
            background: #FFFFFF;
            border: 1px solid #E2E8F0;
            border-radius: 10px;
            padding: 1.1rem 1.25rem;
            box-shadow: 0 2px 6px rgba(0, 0, 0, 0.03);
            transition: transform 0.2s ease, box-shadow 0.2s ease;
        }
        div[data-testid="stMetric"]:hover {
            transform: translateY(-2px);
            box-shadow: 0 6px 16px rgba(0, 0, 0, 0.06);
            border-color: #CBD5E1;
        }
        div[data-testid="stMetricLabel"] {
            font-weight: 600;
            color: #64748B !important;
            font-size: 0.85rem !important;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }
        div[data-testid="stMetricValue"] {
            font-weight: 700;
            color: #0F172A !important;
            font-size: 1.85rem !important;
        }
        .executive-card {
            background-color: #F8FAFC;
            border: 1px solid #E2E8F0;
            border-left: 5px solid #1E3A8A;
            border-radius: 8px;
            padding: 1.25rem 1.5rem;
            margin-bottom: 1.5rem;
            box-shadow: 0 2px 5px rgba(0, 0, 0, 0.02);
        }
        .executive-card h4 {
            color: #1E3A8A;
            margin-top: 0;
            margin-bottom: 0.75rem;
            font-weight: 700;
            font-size: 1.1rem;
        }
        .executive-card ul {
            margin: 0;
            padding-left: 1.25rem;
            color: #334155;
        }
        .executive-card li {
            margin-bottom: 0.5rem;
            line-height: 1.5;
            font-size: 0.95rem;
        }
        .mode-banner {
            background-color: #EFF6FF;
            border: 1px solid #BFDBFE;
            color: #1E40AF;
            padding: 0.6rem 1rem;
            border-radius: 8px;
            font-weight: 600;
            font-size: 0.9rem;
            margin-bottom: 1.2rem;
            display: flex;
            align-items: center;
            justify-content: space-between;
        }
        .stTabs [data-baseweb="tab-list"] {
            gap: 8px;
            border-bottom: 2px solid #E2E8F0;
        }
        .stTabs [data-baseweb="tab"] {
            height: 45px;
            white-space: pre-wrap;
            border-radius: 8px 8px 0 0;
            font-weight: 600;
            color: #64748B;
            padding: 0 18px;
        }
        .stTabs [aria-selected="true"] {
            background-color: #1E3A8A !important;
            color: #FFFFFF !important;
        }
        section[data-testid="stSidebar"] {
            background-color: #F8FAFC;
            border-right: 1px solid #E2E8F0;
        }
        .stDataFrame {
            border-radius: 8px;
            overflow: hidden;
            border: 1px solid #E2E8F0;
        }
    </style>
    """, unsafe_allow_html=True)


def apply_plotly_theme(fig: go.Figure, title: str = "", height: Optional[int] = None) -> go.Figure:
    fig.update_layout(
        title=dict(text=f"<b>{title}</b>" if title else "", font=dict(size=15, color="#0F172A", family="Inter, system-ui, sans-serif")),
        font=dict(family="Inter, system-ui, sans-serif", color="#334155"),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(248,250,252,0.5)",
        margin=dict(l=20, r=20, t=45 if title else 20, b=20),
        xaxis=dict(gridcolor="#E2E8F0", zerolinecolor="#CBD5E1"),
        yaxis=dict(gridcolor="#E2E8F0", zerolinecolor="#CBD5E1"),
    )
    if height:
        fig.update_layout(height=height)
    return fig


def safe_dataframe(df_to_show: pd.DataFrame, show_count: bool = True):
    if df_to_show is not None:
        if show_count and not df_to_show.empty:
            st.caption(f"**Total Records:** {len(df_to_show):,} rows")
        try:
            st.dataframe(df_to_show, use_container_width=True, hide_index=True)
        except TypeError:
            try:
                st.dataframe(df_to_show, use_container_width=True)
            except Exception:
                st.dataframe(df_to_show)


def add_total_count_row(df: pd.DataFrame, label_col: str, count_col: str) -> pd.DataFrame:
    if df.empty or count_col not in df.columns or label_col not in df.columns:
        return df
    total_val = df[count_col].sum()
    total_row = pd.DataFrame([{label_col: "TOTAL SUMMARY", count_col: total_val}])
    return pd.concat([df, total_row], ignore_index=True)


def find_col(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    cols_map = {str(c).strip().lower(): str(c) for c in df.columns}
    for cand in candidates:
        cand_clean = cand.strip().lower()
        if cand_clean in cols_map:
            return cols_map[cand_clean]
    return None


def clean_and_impute_dataframe(df: pd.DataFrame, auto_impute: bool = True) -> Tuple[pd.DataFrame, dict]:
    df = df.copy()

    report = {
        "total_rows": len(df),
        "total_imputed": 0,
        "imputations": {},
        "defaults_applied": {}
    }

    df.columns = df.columns.astype(str).str.strip()
    df = df.loc[:, ~df.columns.duplicated()]

    col_mappings = {
        "Issue Date": ["Issue Date", "Issue date", "Created Date", "Date", "Ticket Date"],
        "Resolution Date": ["Resolution Date", "Resolution date", "Closed Date"],
        "Restoration Date": ["Restoration Date", "Restoration date"],
        "Station ID": ["Station ID", "Station id", "StationID", "Station_ID"],
        "Station Name": ["Station Name", "Station name", "StationName", "Station_Name"],
        "Issue Type": ["Issue Type", "Issue type", "Fault Type", "Category"],
        "Issue Sub-Type": ["Issue Sub-Type", "Issue sub-type", "Sub Category", "Sub Type", "Fault Subtype"],
        "Zone": ["Zone", "Region", "State", "Circle"],
        "Severity": ["Severity", "Priority"],
        "Charger Make": ["Charger Make", "Charger Company", "OEM", "Make", "Vendor"],
        "Status": ["Status", "Ticket Status", "State"],
        "TAT Compliance": ["TAT Compliance", "TAT", "SLA Compliance"]
    }

    for std_name, candidates in col_mappings.items():
        if std_name not in df.columns:
            found = find_col(df, candidates)
            if found and found != std_name:
                df.rename(columns={found: std_name}, inplace=True)

    null_variants = {"nan", "none", "null", "n/a", "na", "-", "", "nat", "undefined"}

    for col in df.select_dtypes(include=["object", "string"]).columns:
        series_str = df[col].astype(str).str.strip()
        is_null_variant = series_str.str.lower().isin(null_variants)
        df[col] = series_str.mask(is_null_variant, np.nan)

    if "Station ID" in df.columns and "Station Name" in df.columns:
        id_missing = df["Station ID"].isna() & df["Station Name"].notna()
        name_missing = df["Station Name"].isna() & df["Station ID"].notna()

        df.loc[id_missing, "Station ID"] = df.loc[id_missing, "Station Name"]
        df.loc[name_missing, "Station Name"] = df.loc[name_missing, "Station ID"]

    if "Status" in df.columns and df["Status"].notna().any():
        df["Status"] = df["Status"].astype(str).str.strip().str.title()

    if "TAT Compliance" in df.columns and df["TAT Compliance"].notna().any():
        tat_upper = df["TAT Compliance"].astype(str).str.upper().str.strip()
        df["TAT Compliance"] = tat_upper.replace({
            "YES": "Yes", "Y": "Yes", "TRUE": "Yes", "1": "Yes", "COMPLIANT": "Yes",
            "NO": "No", "N": "No", "FALSE": "No", "0": "No", "NON-COMPLIANT": "No"
        })

    default_imputations = {
        "Station ID": "UNKNOWN_ID",
        "Station Name": "Unknown Station",
        "Zone": "Unassigned",
        "Charger Make": "Unspecified",
        "Severity": "Unspecified",
        "Issue Type": "Uncategorized",
        "Issue Sub-Type": "Unspecified",
        "Status": "Unknown",
        "TAT Compliance": "N/A"
    }

    if auto_impute:
        for col, default_val in default_imputations.items():
            if col in df.columns:
                null_count = int(df[col].isna().sum())
                if null_count > 0:
                    df[col] = df[col].fillna(default_val)
                    report["imputations"][col] = null_count
                    report["defaults_applied"][col] = default_val
                    report["total_imputed"] += null_count

    for col in DATE_COLS:
        if col in df.columns:
            try:
                df[col] = pd.to_datetime(df[col], errors="coerce", dayfirst=True)
            except Exception:
                df[col] = pd.to_datetime(df[col], errors="coerce")

    if "Issue Date" in df.columns and df["Issue Date"].notna().any():
        issue_dt = df["Issue Date"]
        df["Issue Day"] = issue_dt.dt.date
        df["Issue Month"] = issue_dt.dt.to_period("M").astype(str)
        df["Issue Month"] = df["Issue Month"].replace("NaT", np.nan)

        quarter_str = issue_dt.dt.year.astype(str) + "-Q" + issue_dt.dt.quarter.astype(str)
        df["Quarter"] = quarter_str.where(issue_dt.notna(), np.nan)

    return df, report


@st.cache_data(ttl=300)
def load_data(file, auto_impute: bool = True) -> Tuple[pd.DataFrame, dict]:
    df = pd.read_excel(file)
    cleaned_df, report = clean_and_impute_dataframe(df, auto_impute=auto_impute)
    return cleaned_df, report


def get_repetitive_faults(df: pd.DataFrame) -> pd.DataFrame:
    required = ["Station ID", "Station Name", "Issue Sub-Type"]
    group_cols = [c for c in required if c in df.columns]

    if "Issue Sub-Type" not in df.columns or not ("Station ID" in df.columns or "Station Name" in df.columns):
        return pd.DataFrame(columns=required + ["Occurrences"])

    if df.empty:
        return pd.DataFrame(columns=group_cols + ["Occurrences"])

    counts = (
        df.groupby(group_cols)
        .size()
        .reset_index(name="Occurrences")
    )
    return counts[counts["Occurrences"] >= 2].sort_values("Occurrences", ascending=False)


def get_top_issue_breakdown(df: pd.DataFrame, top_types: int = 5, top_subtypes: int = 5) -> pd.DataFrame:
    if "Issue Type" not in df.columns or "Issue Sub-Type" not in df.columns or df.empty:
        return pd.DataFrame()

    type_counts = df["Issue Type"].dropna().value_counts().head(top_types)
    rows = []
    for issue_type, type_total in type_counts.items():
        sub_counts = (
            df[df["Issue Type"] == issue_type]["Issue Sub-Type"]
            .dropna()
            .value_counts()
            .head(top_subtypes)
        )
        for sub_type, sub_count in sub_counts.items():
            rows.append({
                "Issue Type": issue_type,
                "Issue Type Total": type_total,
                "Issue Sub-Type": sub_type,
                "Sub-Type Count": sub_count,
            })
    return pd.DataFrame(rows)


def get_top_stations(df: pd.DataFrame, n: int = 25) -> pd.DataFrame:
    group_cols = [c for c in ["Station ID", "Station Name"] if c in df.columns]
    if not group_cols or df.empty:
        return pd.DataFrame(columns=["Station Name", "Fault Count"])

    return (
        df.groupby(group_cols)
        .size()
        .reset_index(name="Fault Count")
        .sort_values("Fault Count", ascending=False)
        .head(n)
    )


def get_time_aggregates(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if "Issue Date" not in df.columns or df.empty or df["Issue Date"].isna().all():
        empty = pd.DataFrame(columns=["Time", "Faults"])
        return empty, empty, empty

    valid = df[df["Issue Date"].notna()].copy()

    daily = valid.groupby(valid["Issue Date"].dt.date).size().reset_index(name="Faults")
    daily.columns = ["Time", "Faults"]
    daily["Time"] = daily["Time"].astype(str)

    monthly = valid.groupby(valid["Issue Date"].dt.to_period("M").astype(str)).size().reset_index(name="Faults")
    monthly.columns = ["Time", "Faults"]

    quarterly = valid.groupby(valid["Issue Date"].dt.year.astype(str) + "-Q" + valid["Issue Date"].dt.quarter.astype(str)).size().reset_index(name="Faults")
    quarterly.columns = ["Time", "Faults"]

    return daily, monthly, quarterly


def generate_executive_insights(df: pd.DataFrame, repetitive: pd.DataFrame) -> List[str]:
    insights = []
    total = len(df)
    if total == 0:
        return ["No records match the active filter criteria."]

    if "Status" in df.columns:
        open_count = df["Status"].astype(str).str.strip().str.lower().isin(["open", "pending", "in progress"]).sum()
        open_pct = (open_count / total) * 100
        insights.append(f"**Active Fault Load:** Tracking **{total:,}** total faults, with **{open_count:,} ({open_pct:.1f}%)** currently in open/pending status.")
    else:
        insights.append(f"**Total Volume:** Analyzed **{total:,}** fault reports across selected scope.")

    if "TAT Compliance" in df.columns:
        valid_tat = df["TAT Compliance"].dropna()
        if len(valid_tat) > 0:
            compliant_pct = (valid_tat.astype(str).str.strip().str.lower().isin(["yes", "y", "true", "1", "compliant"]).mean()) * 100
            insights.append(f"**SLA Performance:** Overall Turnaround Time (TAT) compliance stands at **{compliant_pct:.1f}%**.")

    if "Zone" in df.columns:
        top_zone = df["Zone"].mode()
        if not top_zone.empty:
            zone_name = top_zone.iloc[0]
            zone_cnt = (df["Zone"] == zone_name).sum()
            zone_pct = (zone_cnt / total) * 100
            insights.append(f"**Regional Concentration:** **{zone_name}** represents the largest fault share at **{zone_cnt:,} ({zone_pct:.1f}%)** issues.")

    if "Charger Make" in df.columns:
        top_make = df["Charger Make"].mode()
        if not top_make.empty:
            make_name = top_make.iloc[0]
            make_cnt = (df["Charger Make"] == make_name).sum()
            make_pct = (make_cnt / total) * 100
            insights.append(f"**Equipment Distribution:** **{make_name}** equipment accounts for **{make_cnt:,} ({make_pct:.1f}%)** of all reported failures.")

    if not repetitive.empty:
        top_rep = repetitive.iloc[0]
        st_name = top_rep.get("Station Name", top_rep.get("Station ID", "Unknown Station"))
        sub_type = top_rep.get("Issue Sub-Type", "Issue")
        occ = top_rep.get("Occurrences", 0)
        insights.append(f"**Critical Recurring Hotspot:** **{st_name}** has experienced **{occ} repeated instances** of *{sub_type}*.")

    return insights


def generate_executive_excel_report(
    df_all: pd.DataFrame,
    df_filtered: pd.DataFrame,
    repetitive: pd.DataFrame,
    top_stations: pd.DataFrame
) -> bytes:
    output = io.BytesIO()
    wb = openpyxl.Workbook()

    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    title_font = Font(name="Calibri", size=15, bold=True, color="1B365D")
    subtitle_font = Font(name="Calibri", size=10, italic=True, color="64748B")
    section_font = Font(name="Calibri", size=12, bold=True, color="0F172A")
    
    regular_font = Font(name="Calibri", size=10)
    
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1")
    )
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    def style_table(ws, start_row: int, df_to_write: pd.DataFrame) -> int:
        if df_to_write is None or df_to_write.empty:
            return start_row

        headers = list(df_to_write.columns)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_num, value=str(header))
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        ws.row_dimensions[start_row].height = 24

        curr_row = start_row + 1
        for row_idx, (_, row_data) in enumerate(df_to_write.iterrows()):
            ws.row_dimensions[curr_row].height = 20
            for col_num, val in enumerate(row_data, 1):
                cell = ws.cell(row=curr_row, column=col_num)
                if pd.isna(val):
                    cell.value = "N/A"
                elif isinstance(val, (int, np.integer)):
                    cell.value = int(val)
                    cell.number_format = "#,##0"
                elif isinstance(val, (float, np.floating)):
                    cell.value = float(val)
                    cell.number_format = "#,##0.00"
                elif isinstance(val, (pd.Timestamp, datetime)):
                    cell.value = val.strftime("%Y-%m-%d %H:%M")
                else:
                    cell.value = str(val)

                cell.font = regular_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

                if row_idx % 2 == 1:
                    cell.fill = zebra_fill
            curr_row += 1

        return curr_row

    def autofit_cols(ws):
        ws.views.sheetView[0].showGridLines = True
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = 0
            for cell in col:
                if cell.row <= 3:
                    continue
                max_len = max(max_len, len(str(cell.value or "")))
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)

    ws1 = wb.active
    ws1.title = "Executive Overview"
    ws1.cell(row=1, column=1, value="CHARGER FAULT MANAGEMENT SYSTEM").font = title_font
    ws1.cell(row=2, column=1, value=f"Boardroom Executive Summary | Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = subtitle_font

    ws1.cell(row=4, column=1, value="KEY EXECUTIVE PERFORMANCE METRICS").font = section_font

    total_records = len(df_filtered)
    open_count = df_filtered["Status"].astype(str).str.strip().str.lower().isin(["open", "pending", "in progress"]).sum() if "Status" in df_filtered.columns and len(df_filtered) else 0
    rep_pairs = len(repetitive)
    tat_comp = "N/A"
    if "TAT Compliance" in df_filtered.columns and len(df_filtered):
        valid_tat = df_filtered["TAT Compliance"].dropna()
        if len(valid_tat):
            tat_comp = f"{(valid_tat.astype(str).str.strip().str.lower().isin(['yes', 'y', 'true', '1', 'compliant']).mean()) * 100:.1f}%"

    kpi_df = pd.DataFrame([
        {"Performance Indicator": "Total Filtered Fault Volume", "Current Status": f"{total_records:,}"},
        {"Performance Indicator": "Open / Pending Ticket Load", "Current Status": f"{open_count:,}"},
        {"Performance Indicator": "Repetitive Fault Hotspots (2+ Occurrences)", "Current Status": f"{rep_pairs:,}"},
        {"Performance Indicator": "Overall SLA TAT Compliance", "Current Status": tat_comp},
        {"Performance Indicator": "Total Records in Source File", "Current Status": f"{len(df_all):,}"}
    ])
    next_row = style_table(ws1, 5, kpi_df)

    next_row += 2
    ws1.cell(row=next_row, column=1, value="AUTOMATED MEETING EXECUTIVE TAKEAWAYS").font = section_font
    next_row += 1

    insights = generate_executive_insights(df_filtered, repetitive)
    for insight in insights:
        clean_text = insight.replace("**", "").replace("*", "")
        cell = ws1.cell(row=next_row, column=1, value=f"• {clean_text}")
        cell.font = regular_font
        ws1.row_dimensions[next_row].height = 20
        next_row += 1

    autofit_cols(ws1)

    ws2 = wb.create_sheet(title="Time Trends")
    ws2.cell(row=1, column=1, value="TIME-BASED FAULT TREND ANALYSIS").font = title_font
    ws2.cell(row=2, column=1, value="Aggregated breakdown of fault volume across Monthly, Quarterly, and Daily intervals.").font = subtitle_font

    daily, monthly, quarterly = get_time_aggregates(df_filtered)

    ws2.cell(row=4, column=1, value="Month-wise Aggregations").font = section_font
    r = style_table(ws2, 5, add_total_count_row(monthly.sort_values("Time", ascending=False), "Time", "Faults") if not monthly.empty else pd.DataFrame())

    r += 2
    ws2.cell(row=r, column=1, value="Quarter-wise Aggregations").font = section_font
    r = style_table(ws2, r + 1, add_total_count_row(quarterly.sort_values("Time", ascending=False), "Time", "Faults") if not quarterly.empty else pd.DataFrame())

    r += 2
    ws2.cell(row=r, column=1, value="Daily Aggregations (Latest First)").font = section_font
    r = style_table(ws2, r + 1, add_total_count_row(daily.sort_values("Time", ascending=False), "Time", "Faults") if not daily.empty else pd.DataFrame())

    autofit_cols(ws2)

    ws3 = wb.create_sheet(title="Regional & OEM Breakdown")
    ws3.cell(row=1, column=1, value="ZONE & OEM BREAKDOWN SUMMARY").font = title_font

    r = 4
    if "Zone" in df_filtered.columns and not df_filtered.empty:
        ws3.cell(row=r, column=1, value="Fault Distribution by Region (Zone)").font = section_font
        zone_counts = df_filtered["Zone"].dropna().value_counts().reset_index()
        zone_counts.columns = ["Zone", "Fault Count"]
        r = style_table(ws3, r + 1, add_total_count_row(zone_counts, "Zone", "Fault Count"))
        r += 2

    if "Charger Make" in df_filtered.columns and not df_filtered.empty:
        ws3.cell(row=r, column=1, value="Fault Distribution by Charger Company (OEM)").font = section_font
        make_counts = df_filtered["Charger Make"].dropna().value_counts().reset_index()
        make_counts.columns = ["Charger Make", "Fault Count"]
        r = style_table(ws3, r + 1, add_total_count_row(make_counts, "Charger Make", "Fault Count"))
        r += 2

    if "Severity" in df_filtered.columns and not df_filtered.empty:
        ws3.cell(row=r, column=1, value="Severity Distribution").font = section_font
        sev_counts = df_filtered["Severity"].dropna().value_counts().reset_index()
        sev_counts.columns = ["Severity", "Fault Count"]
        r = style_table(ws3, r + 1, add_total_count_row(sev_counts, "Severity", "Fault Count"))

    autofit_cols(ws3)

    ws4 = wb.create_sheet(title="Hotspots & Repetitive")
    ws4.cell(row=1, column=1, value="STATION HOTSPOTS & REPETITIVE FAULTS").font = title_font

    r = 4
    ws4.cell(row=r, column=1, value="Repetitive Fault Pairs (Same Station + Issue Sub-Type >= 2 Times)").font = section_font
    r = style_table(ws4, r + 1, repetitive if not repetitive.empty else pd.DataFrame({"Notice": ["No repetitive faults detected"]}))

    r += 2
    ws4.cell(row=r, column=1, value="Top 25 Stations by Fault Volume").font = section_font
    r = style_table(ws4, r + 1, top_stations if not top_stations.empty else pd.DataFrame({"Notice": ["No station data available"]}))

    autofit_cols(ws4)

    ws5 = wb.create_sheet(title="Filtered Raw Data")
    ws5.cell(row=1, column=1, value="FILTERED RAW FAULT RECORDS").font = title_font
    ws5.cell(row=2, column=1, value=f"Active records matching sidebar filters (Total: {len(df_filtered):,}).").font = subtitle_font

    raw_df_copy = df_filtered.copy()
    for c in DATE_COLS:
        if c in raw_df_copy.columns:
            raw_df_copy[c] = raw_df_copy[c].dt.strftime("%Y-%m-%d %H:%M").fillna("N/A")

    style_table(ws5, 4, raw_df_copy)
    autofit_cols(ws5)

    wb.save(output)
    output.seek(0)
    return output.getvalue()


def apply_filters(df: pd.DataFrame) -> pd.DataFrame:
    st.sidebar.markdown("### Global Filters")
    filtered = df.copy()

    rendered_filters = set()
    filter_configs = [
        ("Quarter", "Quarter"),
        ("Zone", "Zone / Region"),
        ("Severity", "Severity"),
        ("Charger Make", "Charger OEM / Company"),
        ("Issue Type", "Issue Category"),
        ("Status", "Ticket Status")
    ]

    for col, label in filter_configs:
        if col in filtered.columns and col not in rendered_filters:
            rendered_filters.add(col)
            raw_options = filtered[col].dropna().unique()
            unique_opts = list(dict.fromkeys([
                str(x).strip() for x in raw_options
                if pd.notna(x) and str(x).strip() != "" and str(x).strip().lower() not in ["nat", "nan"]
            ]))
            options = sorted(unique_opts, key=lambda s: s.lower())

            if not options:
                continue

            selected = st.sidebar.multiselect(
                label,
                options=options,
                default=options,
                key=f"filter_widget_{col}"
            )

            filtered = filtered[filtered[col].astype(str).str.strip().isin(selected)]

    if "Issue Date" in df.columns and df["Issue Date"].notna().any():
        valid_dates = df["Issue Date"].dropna()
        if not valid_dates.empty:
            min_d = valid_dates.min().date()
            max_d = valid_dates.max().date()

            date_range = st.sidebar.date_input(
                "Issue Date Range",
                value=(min_d, max_d) if min_d != max_d else min_d,
                min_value=min_d,
                max_value=max_d,
                key="filter_widget_issue_date_range"
            )

            if isinstance(date_range, (list, tuple)):
                if len(date_range) == 2:
                    start = pd.Timestamp(date_range[0]).floor("D")
                    end = pd.Timestamp(date_range[1]).replace(hour=23, minute=59, second=59, microsecond=999999)
                    filtered = filtered[(filtered["Issue Date"] >= start) & (filtered["Issue Date"] <= end)]
                elif len(date_range) == 1:
                    start = pd.Timestamp(date_range[0]).floor("D")
                    end = pd.Timestamp(date_range[0]).replace(hour=23, minute=59, second=59, microsecond=999999)
                    filtered = filtered[(filtered["Issue Date"] >= start) & (filtered["Issue Date"] <= end)]

    return filtered


def render_kpis(df: pd.DataFrame, repetitive: pd.DataFrame):
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Total Fault Load", f"{len(df):,}")

    if "Status" in df.columns and len(df):
        open_count = df["Status"].astype(str).str.strip().str.lower().isin(["open", "pending", "in progress"]).sum()
        col2.metric("Open / Pending Tickets", f"{int(open_count):,}")
    else:
        col2.metric("Open / Pending Tickets", "0")

    col3.metric("Repetitive Fault Hotspots", f"{len(repetitive):,}")

    if "TAT Compliance" in df.columns and len(df):
        valid_tat = df["TAT Compliance"].dropna()
        if len(valid_tat):
            compliant = valid_tat.astype(str).str.strip().str.lower().isin(["yes", "y", "true", "1", "compliant"]).mean()
            col4.metric("SLA TAT Compliance", f"{compliant * 100:.1f}%")
        else:
            col4.metric("SLA TAT Compliance", "N/A")
    else:
        col4.metric("SLA TAT Compliance", "N/A")


def main():
    inject_custom_css()

    default_file = DEFAULT_PATH
    if not default_file and os.path.exists("Demo.xlsx"):
        default_file = "Demo.xlsx"

    st.sidebar.title("Settings & Data")

    view_mode = st.sidebar.radio(
        "Dashboard View Mode",
        ["Executive Presentation Mode", "Operational Details Mode"],
        index=0,
        help="Switch view modes: Presentation Mode provides boardroom summaries and high-impact key insights."
    )

    st.sidebar.markdown("---")
    st.sidebar.subheader("Data Source")
    file_path = st.sidebar.text_input("File path on machine", value=default_file, key="input_file_path")
    uploaded = st.sidebar.file_uploader("...or upload Excel file", type=["xlsx"], key="input_file_uploader")

    enable_auto_clean = st.sidebar.checkbox(
        "Auto-clean & Impute Missing Data",
        value=True,
        help="Automatically sanitize whitespaces, normalize column headers, and fill missing fields."
    )

    if st.sidebar.button("Refresh Data Cache", key="btn_refresh"):
        try:
            st.cache_data.clear()
            st.sidebar.success("Data cache cleared!")
        except Exception:
            pass

    if uploaded is not None:
        source = uploaded
    elif file_path:
        path = Path(file_path)
        if not path.exists():
            st.error(f"No file found at path: `{file_path}`")
            return
        source = path
    else:
        st.info("Please specify a valid file path or upload an Excel file to start.")
        return

    try:
        df, cleaning_stats = load_data(source, auto_impute=enable_auto_clean)
    except Exception as e:
        st.error(f"Error reading Excel file: {e}")
        return

    if df.empty:
        st.warning("Loaded file contains no data rows.")
        return

    st.sidebar.markdown("---")
    filtered = apply_filters(df)
    repetitive = get_repetitive_faults(filtered)
    top_stations = get_top_stations(filtered)

    st.sidebar.markdown("---")
    st.sidebar.subheader("Download Executive Report")
    excel_report_bytes = generate_executive_excel_report(df, filtered, repetitive, top_stations)
    
    st.sidebar.download_button(
        label="Download Formatted Excel Report (.xlsx)",
        data=excel_report_bytes,
        file_name=f"Charger_Fault_Executive_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key="btn_download_excel_sidebar"
    )

    csv_bytes = filtered.to_csv(index=False).encode('utf-8')
    st.sidebar.download_button(
        label="Download Raw Filtered Data (.csv)",
        data=csv_bytes,
        file_name=f"Filtered_Faults_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
        mime="text/csv",
        use_container_width=True,
        key="btn_download_csv_sidebar"
    )

    if "Issue Date" in filtered.columns and filtered["Issue Date"].notna().any():
        min_date_str = filtered["Issue Date"].min().strftime("%Y-%m-%d")
        max_date_str = filtered["Issue Date"].max().strftime("%Y-%m-%d")
        date_badge_str = f"Date Scope: {min_date_str} to {max_date_str}"
    else:
        date_badge_str = "Date Scope: Full Historical Dataset"

    st.markdown(f"""
    <div class="header-banner">
        <div style="display: flex; justify-content: space-between; align-items: center;">
            <div>
                <h1 class="header-title">Charger Fault Executive Dashboard</h1>
                <div class="header-subtitle">Boardroom Operations & Reliability Analytics Platform</div>
                <div class="header-badge">Showing {len(filtered):,} of {len(df):,} Total Records | {date_badge_str}</div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    if "Presentation" in view_mode:
        st.markdown("""
        <div class="mode-banner">
            <div><strong>Executive Presentation Mode Active:</strong> Showing high-impact boardroom indicators and strategic takeaways.</div>
        </div>
        """, unsafe_allow_html=True)

    tab_overview, tab_time, tab_zone_oem, tab_repetitive, tab_raw = st.tabs([
        "Executive Overview",
        "Time Series & Trends",
        "Zone & OEM Breakdown",
        "Repetitive & Hotspots",
        "Data & Quality Log"
    ])

    with tab_overview:
        st.subheader("Key Operational Performance Indicators")
        render_kpis(filtered, repetitive)

        st.markdown("<br>", unsafe_allow_html=True)

        insights = generate_executive_insights(filtered, repetitive)
        st.markdown(f"""
        <div class="executive-card">
            <h4>Executive Meeting Takeaways & Key Takeouts</h4>
            <ul>
                {''.join(f'<li>{i}</li>' for i in insights)}
            </ul>
        </div>
        """, unsafe_allow_html=True)

        col_left, col_right = st.columns(2)

        with col_left:
            if "Zone" in filtered.columns and not filtered.empty:
                zone_counts = filtered["Zone"].dropna().value_counts().reset_index()
                zone_counts.columns = ["Zone", "Fault Count"]
                if not zone_counts.empty:
                    fig = px.bar(
                        zone_counts,
                        x="Zone",
                        y="Fault Count",
                        title="Fault Distribution by Region (Zone)",
                        text_auto=True,
                        color_discrete_sequence=[CORPORATE_PALETTE[0]]
                    )
                    apply_plotly_theme(fig, "Fault Distribution by Region (Zone)", height=380)
                    st.plotly_chart(fig, use_container_width=True)

        with col_right:
            if "Charger Make" in filtered.columns and not filtered.empty:
                make_counts = filtered["Charger Make"].dropna().value_counts().reset_index()
                make_counts.columns = ["Charger Make", "Fault Count"]
                if not make_counts.empty:
                    fig = px.pie(
                        make_counts,
                        names="Charger Make",
                        values="Fault Count",
                        title="Fault Share by Charger OEM",
                        hole=0.4,
                        color_discrete_sequence=CORPORATE_PALETTE
                    )
                    apply_plotly_theme(fig, "Fault Share by Charger OEM", height=380)
                    st.plotly_chart(fig, use_container_width=True)

    with tab_time:
        daily, monthly, quarterly = get_time_aggregates(filtered)

        st.subheader("Month-wise Fault Volume")
        if monthly.empty:
            st.info("No Issue Date data available for monthly trend analysis.")
        else:
            fig = px.bar(
                monthly,
                x="Time",
                y="Faults",
                title="Monthly Fault Volume",
                text_auto=True,
                color_discrete_sequence=[CORPORATE_PALETTE[1]]
            )
            apply_plotly_theme(fig, "Monthly Fault Volume", height=350)
            st.plotly_chart(fig, use_container_width=True)
            with st.expander("View Monthly Aggregation Table", expanded=False):
                safe_dataframe(add_total_count_row(monthly.sort_values("Time", ascending=False), "Time", "Faults"))

        st.subheader("Quarter-wise Fault Volume")
        if quarterly.empty:
            st.info("No Issue Date data available for quarterly trend analysis.")
        else:
            fig = px.bar(
                quarterly,
                x="Time",
                y="Faults",
                title="Quarterly Fault Volume",
                text_auto=True,
                color_discrete_sequence=[CORPORATE_PALETTE[2]]
            )
            apply_plotly_theme(fig, "Quarterly Fault Volume", height=350)
            st.plotly_chart(fig, use_container_width=True)
            with st.expander("View Quarterly Aggregation Table", expanded=False):
                safe_dataframe(add_total_count_row(quarterly.sort_values("Time", ascending=False), "Time", "Faults"))

        st.subheader("Daily Fault Volume")
        if daily.empty:
            st.info("No Issue Date data available for daily trend analysis.")
        else:
            fig = px.line(
                daily,
                x="Time",
                y="Faults",
                title="Daily Fault Trend",
                markers=True,
                color_discrete_sequence=[CORPORATE_PALETTE[0]]
            )
            apply_plotly_theme(fig, "Daily Fault Trend", height=350)
            st.plotly_chart(fig, use_container_width=True)
            with st.expander("View Daily Aggregation Table", expanded=False):
                safe_dataframe(add_total_count_row(daily.sort_values("Time", ascending=False), "Time", "Faults"))

    with tab_zone_oem:
        col_a, col_b = st.columns(2)

        with col_a:
            st.subheader("Region (Zone) Breakdown")
            if "Zone" in filtered.columns and not filtered.empty:
                z_counts = filtered["Zone"].dropna().value_counts().reset_index()
                z_counts.columns = ["Zone", "Fault Count"]
                safe_dataframe(add_total_count_row(z_counts, "Zone", "Fault Count"))

        with col_b:
            st.subheader("Charger OEM / Company Breakdown")
            if "Charger Make" in filtered.columns and not filtered.empty:
                m_counts = filtered["Charger Make"].dropna().value_counts().reset_index()
                m_counts.columns = ["Charger Make", "Fault Count"]
                safe_dataframe(add_total_count_row(m_counts, "Charger Make", "Fault Count"))

        st.markdown("<br>", unsafe_allow_html=True)
        c_sev, c_stat = st.columns(2)

        with c_sev:
            st.subheader("Severity Breakdown")
            if "Severity" in filtered.columns and not filtered.empty:
                sev_counts = filtered["Severity"].dropna().value_counts().reset_index()
                sev_counts.columns = ["Severity", "Fault Count"]
                safe_dataframe(add_total_count_row(sev_counts, "Severity", "Fault Count"))

        with c_stat:
            st.subheader("Ticket Status Breakdown")
            if "Status" in filtered.columns and not filtered.empty:
                status_counts = filtered["Status"].dropna().value_counts().reset_index()
                status_counts.columns = ["Status", "Fault Count"]
                safe_dataframe(add_total_count_row(status_counts, "Status", "Fault Count"))

        st.markdown("<br>", unsafe_allow_html=True)
        st.subheader("Top 5 Issue Categories & Sub-Types")
        issue_breakdown = get_top_issue_breakdown(filtered)
        if issue_breakdown.empty:
            st.info("No 'Issue Type' or 'Issue Sub-Type' columns present in dataset.")
        else:
            for issue_type in issue_breakdown["Issue Type"].unique():
                group = issue_breakdown[issue_breakdown["Issue Type"] == issue_type]
                total = group["Issue Type Total"].iloc[0]
                with st.expander(f"{issue_type} — {total} total faults"):
                    fig = px.bar(
                        group,
                        x="Sub-Type Count",
                        y="Issue Sub-Type",
                        orientation="h",
                        title=f"Sub-Type Breakdown for {issue_type}",
                        text_auto=True,
                        color_discrete_sequence=[CORPORATE_PALETTE[4]]
                    )
                    apply_plotly_theme(fig, f"Sub-Type Breakdown for {issue_type}", height=max(250, len(group) * 35))
                    fig.update_layout(yaxis=dict(autorange="reversed", title="Issue Sub-Type"), xaxis_title="Sub-Type Count")
                    st.plotly_chart(fig, use_container_width=True)
                    safe_dataframe(group[["Issue Sub-Type", "Sub-Type Count"]])

    with tab_repetitive:
        st.subheader("Top 25 Hotspot Stations by Fault Volume")
        if top_stations.empty:
            st.info("No station fault data available.")
        else:
            fig = px.bar(
                top_stations,
                x="Fault Count",
                y="Station Name",
                orientation="h",
                title="Top 25 Stations by Total Faults",
                text_auto=True,
                color_discrete_sequence=[CORPORATE_PALETTE[0]]
            )
            apply_plotly_theme(fig, "Top 25 Stations by Total Faults", height=max(400, len(top_stations) * 26))
            fig.update_layout(yaxis=dict(autorange="reversed", title="Station Name"), xaxis_title="Fault Count")
            st.plotly_chart(fig, use_container_width=True)
            with st.expander("View Hotspot Stations Table", expanded=False):
                safe_dataframe(top_stations)

        st.markdown("<br>", unsafe_allow_html=True)
        st.subheader("Repetitive Fault Pairs (Same Station + Issue Sub-Type >= 2 Times)")
        if repetitive.empty:
            st.info("No repetitive faults found for current filter selection.")
        else:
            safe_dataframe(repetitive)

    with tab_raw:
        st.subheader("Data Quality & Imputation Log")
        if cleaning_stats and cleaning_stats.get("total_imputed", 0) > 0:
            st.info(f"Processed **{cleaning_stats['total_rows']:,}** records | Auto-imputed **{cleaning_stats['total_imputed']:,}** missing categorical values.")
            imputations = cleaning_stats.get("imputations", {})
            if imputations:
                imp_rows = [
                    {
                        "Column Name": col,
                        "Missing Values Imputed": count,
                        "Default Value Applied": cleaning_stats.get("defaults_applied", {}).get(col, "N/A")
                    }
                    for col, count in imputations.items() if count > 0
                ]
                if imp_rows:
                    safe_dataframe(pd.DataFrame(imp_rows), show_count=False)
        else:
            st.success("Data quality check complete: 100% complete records (no missing values imputed).")

        st.markdown("<br>", unsafe_allow_html=True)
        st.subheader("Filtered Raw Dataset")
        display_df = filtered.sort_values("Issue Date", ascending=False) if "Issue Date" in filtered.columns else filtered
        safe_dataframe(display_df)


if __name__ == "__main__":
    main()
